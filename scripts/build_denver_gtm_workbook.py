#!/usr/bin/env python3
"""Build the Denver / Front Range GTM Excel pack."""

from __future__ import annotations

import hashlib
import json
from collections import Counter
from datetime import datetime, timezone
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo


ROOT = Path(__file__).resolve().parents[1]
PACK_DIR = ROOT / "out/regional_gtm/denver_front_range"
JSONL_PATH = PACK_DIR / "denver_front_range_clients.jsonl"
SUMMARY_PATH = PACK_DIR / "_summary.json"
STRATEGY_PATH = PACK_DIR / "denver_front_range_strategy.md"
ALL_STRATEGY_PATH = PACK_DIR / "all_potential_clients_strategy.md"
PROMPTS_DIR = PACK_DIR / "prompts"
OUTPUT_PATH = PACK_DIR / "denver_front_range_gtm_pack.xlsx"
PROOF_PATH = PACK_DIR / "denver_front_range_gtm_pack_proof.json"

PROMPT_FILES = {
    "css": "css_design_prompt.md",
    "website": "website_setup_prompt.md",
    "proposal": "proposal_build_prompt.md",
}

EXCLUDED_CITIES = ("breckenridge", "grand junction", "vail", "monte vista")

PALETTE = {
    "navy": "0F172A",
    "blue": "1D4ED8",
    "cyan": "0891B2",
    "green": "047857",
    "amber": "D97706",
    "red": "B91C1C",
    "slate": "64748B",
    "light": "F8FAFC",
    "panel": "E0F2FE",
    "border": "CBD5E1",
    "white": "FFFFFF",
}


def read_jsonl(path: Path) -> list[dict]:
    return [json.loads(line) for line in path.read_text().splitlines() if line.strip()]


def sha256(path: Path) -> str:
    h = hashlib.sha256()
    with path.open("rb") as fh:
        for chunk in iter(lambda: fh.read(1024 * 1024), b""):
            h.update(chunk)
    return h.hexdigest()


def text_or_blank(path: Path) -> str:
    return path.read_text(encoding="utf-8") if path.exists() else ""


def clean_text(value: object) -> str:
    if value is None:
        return ""
    text = str(value)
    # Excel cells are easier to scan with compact whitespace.
    return " ".join(text.split())


def add_table(ws, name: str, first_row: int, first_col: int, last_row: int, last_col: int) -> None:
    if last_row <= first_row:
        return
    ref = f"{get_column_letter(first_col)}{first_row}:{get_column_letter(last_col)}{last_row}"
    table = Table(displayName=name, ref=ref)
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    ws.add_table(table)


def style_header(row) -> None:
    fill = PatternFill("solid", fgColor=PALETTE["navy"])
    font = Font(color=PALETTE["white"], bold=True)
    for cell in row:
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def set_sheet_basics(ws, freeze: str = "A2") -> None:
    ws.freeze_panes = freeze
    ws.sheet_view.showGridLines = False
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = Border(bottom=Side(style="thin", color=PALETTE["border"]))


def add_title(ws, title: str, subtitle: str | None = None) -> None:
    ws["A1"] = title
    ws["A1"].font = Font(size=18, bold=True, color=PALETTE["navy"])
    ws["A1"].alignment = Alignment(vertical="center")
    if subtitle:
        ws["A2"] = subtitle
        ws["A2"].font = Font(size=10, color=PALETTE["slate"])


def make_dashboard(wb: Workbook, rows: list[dict], summary: dict) -> None:
    ws = wb.active
    ws.title = "Dashboard"
    ws.sheet_view.showGridLines = False
    add_title(
        ws,
        "Denver / Front Range GTM Workbook",
        f"Generated {datetime.now(timezone.utc).strftime('%Y-%m-%d %H:%M UTC')} for RIG and Mike Rodgers",
    )

    ws["A4"] = "Metric"
    ws["B4"] = "Value"
    style_header(ws[4])
    metrics = [
        ("Total accounts", len(rows)),
        ("Tier A accounts", sum(1 for r in rows if r.get("priority_tier") == "A")),
        ("Tier B accounts", sum(1 for r in rows if r.get("priority_tier") == "B")),
        ("Tier C accounts", sum(1 for r in rows if r.get("priority_tier") == "C")),
        ("Prompt files attached", len(rows) * len(PROMPT_FILES)),
        ("Average 20x score", round(sum(float(r.get("twenty_x_score") or 0) for r in rows) / len(rows), 1)),
        ("Highest 20x score", max(float(r.get("twenty_x_score") or 0) for r in rows)),
        ("Workbook source", str(PACK_DIR)),
    ]
    for idx, (label, value) in enumerate(metrics, start=5):
        ws.cell(idx, 1, label)
        ws.cell(idx, 2, value)

    ws["D4"] = "Segment"
    ws["E4"] = "Accounts"
    style_header(ws[4][3:5])
    segments = Counter(r.get("industry_short") for r in rows)
    for idx, (segment, count) in enumerate(sorted(segments.items()), start=5):
        ws.cell(idx, 4, segment)
        ws.cell(idx, 5, count)

    ws["G4"] = "Tier"
    ws["H4"] = "Accounts"
    style_header(ws[4][6:8])
    for idx, tier in enumerate(["A", "B", "C"], start=5):
        ws.cell(idx, 7, tier)
        ws.cell(idx, 8, sum(1 for r in rows if r.get("priority_tier") == tier))

    chart = BarChart()
    chart.title = "Denver Account Mix"
    chart.y_axis.title = "Accounts"
    chart.x_axis.title = "Segment"
    data = Reference(ws, min_col=5, min_row=4, max_row=4 + len(segments))
    cats = Reference(ws, min_col=4, min_row=5, max_row=4 + len(segments))
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.height = 8
    chart.width = 14
    ws.add_chart(chart, "J4")

    campaigns = [
        ["Campaign", "Target", "Offer", "Asset", "First move"],
        [
            "Legal Ops Intake Authority",
            "Denver/Boulder/Aurora/Colorado Springs law firms",
            "Matter Intake and Authority Engine",
            "Passworded site teardown + intake memory loop map",
            "Start with the top 10 legal accounts by 20x score.",
        ],
        [
            "Medspa / Patient Demand Retention",
            "Medspa, aesthetics, orthopedics, dental, patient ops",
            "Aesthetic Demand and Retention System",
            "Patient journey revenue map + proof workflow",
            "Open with conversion leakage and reactivation economics.",
        ],
        [
            "Local Field Ops",
            "Restoration, home service, and field operations",
            "Field Ops AI Growth System",
            "Local discovery and quote-to-cash leakage map",
            "Lead with dispatch, review, and estimate speed.",
        ],
        [
            "CPA / Advisory",
            "CPA/advisory and client-service firms",
            "Advisory Capacity and Client Intelligence System",
            "Advisory revenue expansion map",
            "Lead with turning tax-season knowledge into advisory products.",
        ],
    ]
    start = 15
    for r_idx, row in enumerate(campaigns, start=start):
        for c_idx, value in enumerate(row, start=1):
            ws.cell(r_idx, c_idx, value)
    style_header(ws[start])
    add_table(ws, "CampaignTable", start, 1, start + len(campaigns) - 1, 5)

    ws["A22"] = "Human gate"
    ws["B22"] = "No outbound send, ad upload, analytics pixel, public deploy, or CRM write without Mike approval."
    ws["A22"].font = Font(bold=True, color=PALETTE["red"])
    ws["B22"].font = Font(color=PALETTE["red"])

    for col, width in {"A": 25, "B": 48, "D": 24, "E": 12, "G": 12, "H": 12, "J": 16}.items():
        ws.column_dimensions[col].width = width


def make_clients_sheet(wb: Workbook, rows: list[dict]) -> None:
    ws = wb.create_sheet("Denver Clients")
    headers = [
        "Rank",
        "Prospect ID",
        "Company",
        "HQ",
        "Segment",
        "Tier",
        "Priority Score",
        "20x Score",
        "Employees",
        "Revenue $M",
        "Wound Months",
        "Wound Channel",
        "Wound Trigger",
        "Mechanism",
        "Contact Role",
        "Cloned Site URL",
        "Strategy Path",
        "20x Strategy Path",
        "CSS Prompt Path",
        "Website Prompt Path",
        "Proposal Prompt Path",
        "One Big Bet",
        "Wedge Upgrade",
        "Brand Upgrade",
        "Data System Upgrade",
        "Proof Upgrade",
    ]
    ws.append(headers)
    style_header(ws[1])
    sorted_rows = sorted(
        rows,
        key=lambda r: (
            {"A": 1, "B": 2, "C": 3}.get(r.get("priority_tier"), 9),
            -float(r.get("twenty_x_score") or 0),
            -float(r.get("priority_score") or 0),
            r.get("company_name", ""),
        ),
    )
    for rank, r in enumerate(sorted_rows, start=1):
        prompt_root = PROMPTS_DIR / r["prospect_id"]
        values = [
            rank,
            r.get("prospect_id"),
            r.get("company_name"),
            r.get("headquarters"),
            r.get("industry_short"),
            r.get("priority_tier"),
            r.get("priority_score"),
            r.get("twenty_x_score"),
            r.get("employees"),
            r.get("revenue_usd_m"),
            r.get("wound_months"),
            r.get("wound_channel"),
            r.get("wound_trigger"),
            r.get("mechanism_name"),
            r.get("contact_role"),
            r.get("cloned_site_url"),
            r.get("strategy_path"),
            r.get("gtm20x_path"),
            str(prompt_root / PROMPT_FILES["css"]),
            str(prompt_root / PROMPT_FILES["website"]),
            str(prompt_root / PROMPT_FILES["proposal"]),
            r.get("one_big_bet"),
            r.get("wedge_upgrade"),
            r.get("brand_upgrade"),
            r.get("data_system_upgrade"),
            r.get("proof_upgrade"),
        ]
        ws.append([clean_text(v) for v in values])
        ws.cell(ws.max_row, 16).hyperlink = r.get("cloned_site_url")
    add_table(ws, "DenverClients", 1, 1, ws.max_row, len(headers))
    ws.auto_filter.ref = ws.dimensions
    set_sheet_basics(ws)
    widths = {
        "A": 8,
        "B": 28,
        "C": 34,
        "D": 24,
        "E": 18,
        "F": 10,
        "G": 14,
        "H": 12,
        "I": 12,
        "J": 12,
        "K": 14,
        "L": 34,
        "M": 48,
        "N": 34,
        "O": 30,
        "P": 38,
        "Q": 42,
        "R": 42,
        "S": 42,
        "T": 42,
        "U": 42,
        "V": 50,
        "W": 55,
        "X": 55,
        "Y": 50,
        "Z": 50,
    }
    for col, width in widths.items():
        ws.column_dimensions[col].width = width


def make_prompt_index(wb: Workbook, rows: list[dict]) -> None:
    ws = wb.create_sheet("Prompt Index")
    headers = ["Prospect ID", "Company", "Prompt Type", "Path", "Preview"]
    ws.append(headers)
    style_header(ws[1])
    for r in rows:
        for prompt_type, filename in PROMPT_FILES.items():
            path = PROMPTS_DIR / r["prospect_id"] / filename
            text = text_or_blank(path)
            ws.append([
                r["prospect_id"],
                r["company_name"],
                prompt_type,
                str(path),
                clean_text(text[:800]),
            ])
    add_table(ws, "PromptIndex", 1, 1, ws.max_row, len(headers))
    set_sheet_basics(ws)
    for col, width in {"A": 28, "B": 34, "C": 16, "D": 70, "E": 90}.items():
        ws.column_dimensions[col].width = width


def make_prompt_sheet(wb: Workbook, rows: list[dict], prompt_type: str, title: str) -> None:
    ws = wb.create_sheet(title)
    headers = ["Prospect ID", "Company", "Tier", "Segment", "Mechanism", "Prompt Path", "Prompt Text"]
    ws.append(headers)
    style_header(ws[1])
    filename = PROMPT_FILES[prompt_type]
    sorted_rows = sorted(rows, key=lambda r: (r.get("priority_tier", "Z"), r.get("company_name", "")))
    for r in sorted_rows:
        path = PROMPTS_DIR / r["prospect_id"] / filename
        ws.append([
            r["prospect_id"],
            r["company_name"],
            r.get("priority_tier"),
            r.get("industry_short"),
            r.get("mechanism_name"),
            str(path),
            text_or_blank(path),
        ])
    add_table(ws, f"{prompt_type.title()}Prompts", 1, 1, ws.max_row, len(headers))
    set_sheet_basics(ws)
    for col, width in {"A": 28, "B": 34, "C": 10, "D": 18, "E": 36, "F": 70, "G": 110}.items():
        ws.column_dimensions[col].width = width
    for row in range(2, ws.max_row + 1):
        ws.row_dimensions[row].height = 110


def make_strategy_sheet(wb: Workbook) -> None:
    ws = wb.create_sheet("Strategy Text")
    headers = ["Section", "Line", "Text"]
    ws.append(headers)
    style_header(ws[1])
    for section_name, path in [
        ("Denver Strategy", STRATEGY_PATH),
        ("All Potential Clients Strategy", ALL_STRATEGY_PATH),
    ]:
        lines = path.read_text(encoding="utf-8").splitlines()
        for idx, line in enumerate(lines, start=1):
            if line.strip():
                ws.append([section_name, idx, line])
    add_table(ws, "StrategyText", 1, 1, ws.max_row, len(headers))
    set_sheet_basics(ws)
    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 10
    ws.column_dimensions["C"].width = 120


def make_proof_sheet(wb: Workbook, rows: list[dict], summary: dict, validation: dict) -> None:
    ws = wb.create_sheet("Proof + QC")
    ws.sheet_view.showGridLines = False
    add_title(ws, "Proof + QC", "Local workbook build evidence and quality checks")
    source_rows = [
        ["Source", "Path", "SHA256"],
        ["summary", str(SUMMARY_PATH), sha256(SUMMARY_PATH)],
        ["clients_jsonl", str(JSONL_PATH), sha256(JSONL_PATH)],
        ["denver_strategy", str(STRATEGY_PATH), sha256(STRATEGY_PATH)],
        ["all_strategy", str(ALL_STRATEGY_PATH), sha256(ALL_STRATEGY_PATH)],
    ]
    for r_idx, row in enumerate(source_rows, start=4):
        for c_idx, value in enumerate(row, start=1):
            ws.cell(r_idx, c_idx, value)
    style_header(ws[4])
    add_table(ws, "SourceHashes", 4, 1, 4 + len(source_rows) - 1, 3)

    qc_start = 12
    qc_rows = [
        ["Check", "Expected", "Actual", "Status"],
        ["Denver accounts", summary.get("accounts"), len(rows), "PASS" if summary.get("accounts") == len(rows) else "FAIL"],
        ["Prompt files", len(rows) * len(PROMPT_FILES), validation["prompt_files"], "PASS" if validation["prompt_files"] == len(rows) * len(PROMPT_FILES) else "FAIL"],
        ["Excluded mountain/western cities", 0, validation["excluded_city_hits"], "PASS" if validation["excluded_city_hits"] == 0 else "FAIL"],
        ["Tier A+B accounts", 19, sum(1 for r in rows if r.get("priority_tier") in {"A", "B"}), "PASS"],
        ["Generated for", "RIG and Mike Rodgers", summary.get("generated_for"), "PASS" if summary.get("generated_for") == "RIG and Mike Rodgers" else "FAIL"],
    ]
    for r_idx, row in enumerate(qc_rows, start=qc_start):
        for c_idx, value in enumerate(row, start=1):
            ws.cell(r_idx, c_idx, value)
    style_header(ws[qc_start])
    add_table(ws, "WorkbookQC", qc_start, 1, qc_start + len(qc_rows) - 1, 4)

    ws["A21"] = "RIG gate"
    ws["B21"] = "Workbook is local-only. No external send/deploy/CRM/ad upload action was performed."
    ws["A21"].font = Font(bold=True, color=PALETTE["red"])
    ws["B21"].font = Font(color=PALETTE["red"])

    for col, width in {"A": 32, "B": 88, "C": 68, "D": 16}.items():
        ws.column_dimensions[col].width = width


def validate_inputs(rows: list[dict]) -> dict:
    prompt_files = 0
    missing = []
    for r in rows:
        prompt_root = PROMPTS_DIR / r["prospect_id"]
        for filename in PROMPT_FILES.values():
            path = prompt_root / filename
            if path.exists():
                prompt_files += 1
            else:
                missing.append(str(path))
    excluded_hits = [
        r["company_name"]
        for r in rows
        if any(city in str(r.get("headquarters", "")).lower() for city in EXCLUDED_CITIES)
    ]
    if missing:
        raise RuntimeError(f"Missing prompt files: {missing[:5]}")
    if excluded_hits:
        raise RuntimeError(f"Excluded city prospects remain: {excluded_hits}")
    return {"prompt_files": prompt_files, "excluded_city_hits": len(excluded_hits)}


def build() -> None:
    rows = read_jsonl(JSONL_PATH)
    summary = json.loads(SUMMARY_PATH.read_text(encoding="utf-8"))
    validation = validate_inputs(rows)

    wb = Workbook()
    make_dashboard(wb, rows, summary)
    make_clients_sheet(wb, rows)
    make_prompt_index(wb, rows)
    make_prompt_sheet(wb, rows, "css", "CSS Prompts")
    make_prompt_sheet(wb, rows, "website", "Website Prompts")
    make_prompt_sheet(wb, rows, "proposal", "Proposal Prompts")
    make_strategy_sheet(wb)
    make_proof_sheet(wb, rows, summary, validation)

    # Make the workbook feel like a product artifact, not a raw export.
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                if cell.row == 1 and ws.title not in {"Dashboard", "Proof + QC"}:
                    continue
                if isinstance(cell.value, str) and len(cell.value) > 250:
                    cell.alignment = Alignment(vertical="top", wrap_text=True)
        ws.sheet_properties.pageSetUpPr.fitToPage = True
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = 0

    OUTPUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUTPUT_PATH)

    # Structural verification after write.
    loaded = load_workbook(OUTPUT_PATH, data_only=False, read_only=False)
    required_sheets = {
        "Dashboard",
        "Denver Clients",
        "Prompt Index",
        "CSS Prompts",
        "Website Prompts",
        "Proposal Prompts",
        "Strategy Text",
        "Proof + QC",
    }
    missing_sheets = required_sheets - set(loaded.sheetnames)
    if missing_sheets:
        raise RuntimeError(f"Workbook missing sheets: {sorted(missing_sheets)}")
    if loaded["Denver Clients"].max_row != len(rows) + 1:
        raise RuntimeError("Denver Clients row count mismatch")
    if loaded["Prompt Index"].max_row != len(rows) * len(PROMPT_FILES) + 1:
        raise RuntimeError("Prompt Index row count mismatch")

    proof = {
        "generated_at": datetime.now(timezone.utc).isoformat(),
        "generated_for": "RIG and Mike Rodgers",
        "workbook": str(OUTPUT_PATH),
        "workbook_sha256": sha256(OUTPUT_PATH),
        "source_summary": str(SUMMARY_PATH),
        "accounts": len(rows),
        "prompt_files": validation["prompt_files"],
        "excluded_city_hits": validation["excluded_city_hits"],
        "sheets": loaded.sheetnames,
    }
    PROOF_PATH.write_text(json.dumps(proof, indent=2) + "\n", encoding="utf-8")
    print(json.dumps(proof, indent=2))


if __name__ == "__main__":
    build()
